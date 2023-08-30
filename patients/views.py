# Django imports
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from django.views import View

# External library for Excel export
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

# Project-specific imports
from .models import Patient
from .forms import RegisterPatientForm, PatientFilterForm
from django.contrib import messages
from django.http import JsonResponse


def add_new_patient(request):
    if request.method == 'POST':
        form = RegisterPatientForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()  # This will save the user and the associated patient
            
            # Display a success message after adding the patient.
            messages.success(request, f'Patient {user.username} has been added successfully!')
            
            return redirect('patient_list')  # Redirect to patient list view
    else:
        form = RegisterPatientForm()

    context = {'form': form}
    return render(request, 'add_new_patient.html', context)


def patient_list(request):
    # Base patient query. Providing a default ordering mechanism.
    patients = Patient.objects.all().order_by('-date_added')

    # Initialize the filter form
    form = PatientFilterForm(request.GET)

    # Filtering patients using the form's logic
    if form.is_valid():
        patients = form.filter_queryset(patients)



    # In your patient_list view
    latest = request.GET.get('latest', None)
    if latest is not None:
     try:
        latest = int(latest)
        patients = patients[:latest]
     except ValueError:
        # Handle invalid input
        pass


    # Exporting patients to Excel
    if 'export_excel' in request.GET:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="patient_list.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Patient List"

        # Write headers
        headers = ["ID", "Name", "Address", "Gender", "Date of Birth", "Phone", "Location"]
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header_title
            ws.column_dimensions[col_letter].width = 15

        # Write data
        for row_num, patient in enumerate(patients, 2):
            ws[f"A{row_num}"] = patient.patient_id
            ws[f"B{row_num}"] = f"{patient.user.first_name} {patient.user.last_name}"
            ws[f"C{row_num}"] = patient.address
            ws[f"D{row_num}"] = patient.gender
            ws[f"E{row_num}"] = patient.date_of_birth
            ws[f"F{row_num}"] = patient.phone
            ws[f"G{row_num}"] = patient.location.name

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        return response

    # Pagination
    paginator = Paginator(patients, 10)  # Show 10 patients per page
    page_number = request.GET.get('page')
    patients_page = paginator.get_page(page_number)

    

    context = {
        'patients': patients_page,
        'form': form,
    }
    return render(request, 'patient_list.html', context)



class PatientProfileView(View):
    def get(self, request, patient_id):
        patient = get_object_or_404(Patient, id=patient_id)
        context = {'patient': patient}
        return render(request, 'patient_profile.html', context)


def get_patient_details(request, patient_id):
    try:
        patient = Patient.objects.get(id=patient_id)
        data = {
            "id": patient.id,
            "username": patient.username,
            "phone": patient.phone,
            "profile_photo": str(patient.profile_photo.url) if patient.profile_photo else None,
        }
        return JsonResponse(data)
    except Patient.DoesNotExist:
        return JsonResponse({"error": "Patient not found"}, status=404)



def update_patient(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    user = patient.user
    
    if request.method == 'POST':
        form = RegisterPatientForm(request.POST, request.FILES, instance=patient, initial={'email': user.email})
        if form.is_valid():
            user = form.save()  # This updates the user and associated patient
            messages.success(request, f'Patient {user.username} has been updated successfully!')
            return redirect('patient_profile', patient_id=patient.id)
    else:
        form_data = {field.name: getattr(patient, field.name) for field in Patient._meta.fields if field.name in RegisterPatientForm.Meta.fields}
        form_data['email'] = user.email
        form = RegisterPatientForm(initial=form_data)
    
    context = {'form': form, 'patient': patient}
    return render(request, 'update_patient.html', context)


def delete_patient(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    
    if request.method == 'POST':
        patient.delete()
        return redirect('patients:patient_list')  # Redirect to the patient list after deletion

    return render(request, 'delete_patient.html', {'patient': patient})

